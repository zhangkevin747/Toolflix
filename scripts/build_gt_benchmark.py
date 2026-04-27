"""
Ground-truth GAIA benchmark v2.

Every expected_answer is a body-content substring that is NOT in the task
text, NOT in the URL, and NOT in the artifact path. For web-based
categories (wikipedia, fetch) we fetch each URL once and verify the
expected substring is present in the rendered body. For search, we use
factual queries whose canonical answer is not in the query.

This closes the string-match leak that let trivial tool echoes pass in v1.
"""
import json
import re
import time
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from pypdf import PdfReader
import docx


ROOT = Path(__file__).parent.parent
FILES_DIR = ROOT / "data" / "gaia_files"
PDFS_DIR = ROOT / "data" / "pdfs"

HEADERS = {"User-Agent": "Mozilla/5.0 (Toolflix-benchmark-builder)"}


def fetch_body_text(url: str, cache: dict) -> str:
    if url in cache:
        return cache[url]
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"  ! fetch failed: {url}  ({e})")
        cache[url] = ""
        return ""
    soup = BeautifulSoup(r.text, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)
    cache[url] = text
    return text


def assert_not_leaking(task: str, expected: str, artifact: str = ""):
    """Ensure expected_answer is not a substring of task text or artifact path."""
    t_low = task.lower()
    a_low = str(artifact).lower()
    e_low = expected.lower()
    assert e_low not in t_low, f"expected '{expected}' leaks in task: {task!r}"
    assert e_low not in a_low, f"expected '{expected}' leaks in artifact: {artifact!r}"


# ======================================================================
# Wikipedia — 16 tasks
# Pick an entity-internal body fact that is not the entity name and not
# in the task text.
# ======================================================================

WIKI_ENTITIES = [
    # (entity_name, expected_body_fact, verification_url)
    ("Mercedes Sosa",                      "Argentine",          "https://en.wikipedia.org/wiki/Mercedes_Sosa"),
    ("the Moon",                           "natural satellite",  "https://en.wikipedia.org/wiki/Moon"),
    ("Giganotosaurus",                     "Cretaceous",         "https://en.wikipedia.org/wiki/Giganotosaurus"),
    ("Eliud Kipchoge",                     "Kenyan",             "https://en.wikipedia.org/wiki/Eliud_Kipchoge"),
    ("The Lord of the Rings",              "Tolkien",            "https://en.wikipedia.org/wiki/The_Lord_of_the_Rings"),
    ("Antidisestablishmentarianism",       "Church of England",  "https://en.wikipedia.org/wiki/Antidisestablishmentarianism"),
    ("Carl Nebel",                         "lithographer",       "https://en.wikipedia.org/wiki/Carl_Nebel"),
    ("the commutative property",           "binary operation",   "https://en.wikipedia.org/wiki/Commutative_property"),
    ("Albert Einstein",                    "1879",               "https://en.wikipedia.org/wiki/Albert_Einstein"),
    ("Marie Curie",                        "radioactivity",      "https://en.wikipedia.org/wiki/Marie_Curie"),
    ("Python (programming language)",      "van Rossum",         "https://en.wikipedia.org/wiki/Python_(programming_language)"),
    ("the Eiffel Tower",                   "1889",               "https://en.wikipedia.org/wiki/Eiffel_Tower"),
    ("World War II",                       "Allies",             "https://en.wikipedia.org/wiki/World_War_II"),
    ("Mount Everest",                      "8,848",              "https://en.wikipedia.org/wiki/Mount_Everest"),
    ("the Pacific Ocean",                  "largest",            "https://en.wikipedia.org/wiki/Pacific_Ocean"),
    ("William Shakespeare",                "Stratford",          "https://en.wikipedia.org/wiki/William_Shakespeare"),
]


def build_wikipedia_tasks(cache):
    phrasings = [
        "Find the Wikipedia page for {e} and return its content",
        "Look up {e} on Wikipedia and show me what the article says",
        "Get the Wikipedia article about {e}",
        "Fetch the Wikipedia entry for {e}",
    ]
    out = []
    for entity, expected, url in WIKI_ENTITIES:
        body = fetch_body_text(url, cache)
        assert expected.lower() in body.lower(), \
            f"'{expected}' not found in body of {url}"
        task = phrasings[len(out) % len(phrasings)].format(e=entity)
        assert_not_leaking(task, expected, entity)
        out.append({
            "task": task,
            "expected_answer": expected,
            "expected_category": "wikipedia",
            "id": f"wikipedia-{len(out):02d}",
            "artifact": entity,
        })
    return out


# ======================================================================
# Fetch — 18 tasks
# Use a mix of URLs. Expected answer must be in the body, not in the URL
# and not in the task text.
# ======================================================================

FETCH_SPECS = [
    # (url, body_fact_expected)
    # Wikipedia pages — same structural idea as wiki tasks but as direct URL fetch
    ("https://en.wikipedia.org/wiki/Moon",              "natural satellite"),
    ("https://en.wikipedia.org/wiki/Giganotosaurus",    "Cretaceous"),
    ("https://en.wikipedia.org/wiki/Legume",            "Fabaceae"),
    ("https://en.wikipedia.org/wiki/Mercedes_Sosa",     "Argentine"),
    ("https://example.com/",                            "documentation"),
    ("https://www.python.org/",                         "PyCon"),
    ("https://en.wikipedia.org/wiki/Albert_Einstein",   "1879"),
    ("https://en.wikipedia.org/wiki/Marie_Curie",       "radioactivity"),
    ("https://en.wikipedia.org/wiki/World_War_II",      "Allies"),
    ("https://en.wikipedia.org/wiki/William_Shakespeare","Stratford"),
    ("https://en.wikipedia.org/wiki/Mount_Everest",     "8,848"),
    ("https://en.wikipedia.org/wiki/Pacific_Ocean",     "largest"),
    ("https://en.wikipedia.org/wiki/Eiffel_Tower",      "1889"),
    ("https://en.wikipedia.org/wiki/NASA",              "aeronautics"),
    ("https://en.wikipedia.org/wiki/Apollo_11",         "Armstrong"),
    ("https://en.wikipedia.org/wiki/DNA",               "nucleotide"),
    ("https://en.wikipedia.org/wiki/Photosynthesis",    "chlorophyll"),
    ("https://en.wikipedia.org/wiki/Internet",          "ARPANET"),
]


def build_fetch_tasks(cache):
    phrasings = [
        "Fetch the content at {url} and return what's on the page",
        "Download the page {url}",
        "Retrieve the HTML at {url}",
        "Get the webpage at {url}",
    ]
    out = []
    for url, expected in FETCH_SPECS:
        body = fetch_body_text(url, cache)
        assert expected.lower() in body.lower(), \
            f"'{expected}' not found in body of {url}"
        task = phrasings[len(out) % len(phrasings)].format(url=url)
        assert_not_leaking(task, expected, url)
        out.append({
            "task": task,
            "expected_answer": expected,
            "expected_category": "fetch",
            "id": f"fetch-{len(out):02d}",
            "artifact": url,
        })
    return out


# ======================================================================
# Search — 18 tasks
# Factual queries. Expected answer is a fact that search results return
# but is NOT in the query text.
# ======================================================================

SEARCH_SPECS = [
    # Factual queries. Expected answer must NOT be in the query and must be
    # long/distinctive enough to avoid spurious matches in tool output.
    ("capital city of Iceland",              "Reykjav"),                # Reykjavík
    ("founder of Tesla Motors",              "Musk"),
    ("year the French Revolution started",   "1789"),                   # distinctive year, not in query
    ("speed of light in vacuum",             "299,792"),                # distinctive number
    ("author of To Kill a Mockingbird",      "Harper"),
    ("year the Berlin Wall fell",            "1989"),
    ("longest river in Africa",              "Nile"),
    ("largest planet in the solar system",   "Jupiter"),
    ("capital of Australia",                 "Canberra"),
    ("currency of Japan",                    "yen"),
    ("discovery year of penicillin",         "1928"),
    ("first person to walk on the moon",     "Armstrong"),
    ("tallest mountain in North America",    "Denali"),
    ("creator of the Linux kernel",          "Torvalds"),
    ("deepest trench in the ocean",          "Mariana"),
    ("inventor of the telephone",            "Bell"),
    ("home city of the Louvre museum",       "Paris"),
    ("painter of The Starry Night",          "Gogh"),                   # Van Gogh
]


def build_search_tasks():
    phrasings = [
        'Search the web for the {q} and return the results',
        'Google the {q} and show me what comes up',
        'Run a web search to find the {q}',
        'Find search results for the {q}',
    ]
    out = []
    for q, expected in SEARCH_SPECS:
        task = phrasings[len(out) % len(phrasings)].format(q=q)
        assert_not_leaking(task, expected, q)
        out.append({
            "task": task,
            "expected_answer": expected,
            "expected_category": "search",
            "id": f"search-{len(out):02d}",
            "artifact": q,
        })
    return out


# ======================================================================
# Excel — 16 tasks (unchanged — these are already clean).
# ======================================================================

EXCEL_SPECS = [
    ("076c8171-9b3b-49b9-a477-244d2a532826.xlsx", ["Rainforest Bistro", "Panorama Outfitters"]),
    ("32102e3e-d12a-4209-9163-7b3a104efe5d.xlsx", ["Flop Video Rental Store", "Time-Parking 2"]),
    ("4d0aa727-86b1-406b-9b33-f870dd14a4a5.xlsx", ["Sunset Picnic Trip", "Static Display"]),
    ("4d51c4bf-4b0e-4f3d-897b-3f6687a7d9f2.xlsx", ["Halpert", "Begonia Drive"]),
    ("54612da3-fd56-4941-80f4-5eb82330de25.xlsx", ["Operational", "Excursion/Location"]),
    ("7bd855d8-463d-4ed5-93ca-5fe35145f733.xlsx", ["Pinebrook", "Wharvton"]),
    ("7cc4acfa-63fd-4acc-a1a1-e8e529e0a97f.xlsx", ["Sagrada", "Marztep"]),
    ("edd4d4f2-1a58-45c4-b038-67337af4e029.xlsx", ["Main Lawn", "Sunset Picnic Trip"]),
]

EXCEL_PHRASINGS = [
    "Read the Excel file at data/gaia_files/{fname} and return its contents",
    "Open the spreadsheet data/gaia_files/{fname} and show me what's inside",
    "Parse data/gaia_files/{fname} and dump the cell values",
    "Get the data from the xlsx file data/gaia_files/{fname}",
]


def read_xlsx_text(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    chunks = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(max_row=80, values_only=True):
            for c in row:
                if c is not None:
                    chunks.append(str(c))
    return "\n".join(chunks)


def build_excel_tasks():
    out = []
    for fname, strings in EXCEL_SPECS:
        text = read_xlsx_text(FILES_DIR / fname)
        for expected in strings:
            assert expected in text, f"{fname}: '{expected}' not in xlsx"
            task = EXCEL_PHRASINGS[len(out) % len(EXCEL_PHRASINGS)].format(fname=fname)
            # xlsx file names are UUIDs so no leak risk, but assert anyway
            assert_not_leaking(task, expected, fname)
            out.append({
                "task": task,
                "expected_answer": expected,
                "expected_category": "excel",
                "id": f"excel-{len(out):02d}",
                "artifact": fname.split(".")[0],
            })
    return out


# ======================================================================
# PDF — expanded to more artifacts (was only 3; now 8+).
# ======================================================================

PDF_SPECS = [
    # GAIA-distributed PDFs
    ("gaia", "e9a2c537-8232-4c3f-85b0-b52de6bcba99.pdf",
     ["The Very Hungry Caterpillar", "Scribe County Public Library", "Eric Carle"]),
    ("gaia", "366e2f2b-8632-4ef2-81eb-bc3877489217.pdf",
     ["Neptune's Palace", "Sea Escape Inn", "Admiral Sturgeon"]),
    ("gaia", "67e8878b-5cef-4375-804e-e6291fdbe78a.pdf",
     ["Laughing Gull", "Loach Towers"]),
    # Research papers (prose PDFs we downloaded)
    ("pdfs", "attention.pdf",            ["self-attention"]),
    ("pdfs", "1810.04805.pdf",           ["bidirectional"]),
    ("pdfs", "1512.03385.pdf",           ["residual"]),
    ("pdfs", "1606.07792.pdf",           ["Recommender"]),
    ("pdfs", "2201.11903.pdf",           ["reasoning"]),
    ("pdfs", "2302.13971.pdf",           ["tokens"]),
]


def read_pdf_text(path):
    r = PdfReader(str(path))
    return "\n".join((p.extract_text() or "") for p in r.pages[:3])


PDF_PHRASINGS = [
    "Read the PDF at data/{dir}/{fname} and return its contents",
    "Extract the text from data/{dir}/{fname}",
    "Parse the PDF document at data/{dir}/{fname}",
    "Open data/{dir}/{fname} and show the text inside",
]


def build_pdf_tasks():
    out = []
    for dir_key, fname, strings in PDF_SPECS:
        dir_path = FILES_DIR if dir_key == "gaia" else PDFS_DIR
        text = read_pdf_text(dir_path / fname)
        for expected in strings:
            # Case-insensitive presence check
            assert expected.lower() in text.lower(), \
                f"{fname}: '{expected}' not in pdf body"
            task = PDF_PHRASINGS[len(out) % len(PDF_PHRASINGS)].format(dir=f"{'gaia_files' if dir_key == 'gaia' else 'pdfs'}", fname=fname)
            assert_not_leaking(task, expected, fname)
            out.append({
                "task": task,
                "expected_answer": expected,
                "expected_category": "pdf",
                "id": f"pdf-{len(out):02d}",
                "artifact": fname.split(".")[0],
            })
    return out


# ======================================================================
# Filesystem — keep existing set (clean, just reuse files).
# ======================================================================

def build_filesystem_tasks():
    FIX = ROOT / "data" / "fixtures"
    docx_text = "\n".join(p.text for p in docx.Document(
        FILES_DIR / "cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx").paragraphs)
    py_text = (FILES_DIR / "f918266a-b3e0-4914-865d-4faa564f1aef.py").read_text()
    cfg_text = (FIX / "config.json").read_text()
    readme_text = (FIX / "README.md").read_text()
    csv_text = (FIX / "data.csv").read_text()

    specs = [
        # docx
        ("data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", "Rebecca",
         "Read the file at data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", docx_text),
        ("data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", "Harry",
         "Open data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx and show the text", docx_text),
        ("data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", "Micah",
         "What does the file data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx say?", docx_text),
        ("data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", "Georgette",
         "Fetch the contents of data/gaia_files/cffe0e32-c9a6-4c52-9877-78ceb4aaa9fb.docx", docx_text),
        # py
        ("data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", "UhOh",
         "Read the Python file at data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", py_text),
        ("data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", "randint",
         "Show the contents of data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", py_text),
        ("data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", "class Hmm",
         "Open the file data/gaia_files/f918266a-b3e0-4914-865d-4faa564f1aef.py", py_text),
        # config.json fixture
        ("data/fixtures/config.json", "toolflix_db",
         "Read the file at data/fixtures/config.json", cfg_text),
        ("data/fixtures/config.json", "5432",
         "Open data/fixtures/config.json and show the contents", cfg_text),
        # README fixture
        ("data/fixtures/README.md", "report generation",
         "Read the file at data/fixtures/README.md", readme_text),
        ("data/fixtures/README.md", "API integration",
         "Show me what's in data/fixtures/README.md", readme_text),
        # CSV fixture
        ("data/fixtures/data.csv", "gamma",
         "Read the file at data/fixtures/data.csv", csv_text),
        ("data/fixtures/data.csv", "kappa",
         "Open data/fixtures/data.csv and show the contents", csv_text),
    ]
    out = []
    for fname, expected, task, content in specs:
        assert expected in content, f"{fname}: {expected} not in content"
        assert_not_leaking(task, expected, fname)
        out.append({
            "task": task,
            "expected_answer": expected,
            "expected_category": "filesystem",
            "id": f"filesystem-{len(out):02d}",
            "artifact": fname.split(".")[0],
        })
    return out


# ======================================================================
# arXiv — keep existing set (clean).
# ======================================================================

ARXIV_PAPERS = [
    ("1706.03762", "self-attention"),
    ("1810.04805", "bidirectional"),
    ("2005.14165", "few-shot"),
    ("1512.03385", "residual"),
    ("2106.09685", "low-rank"),
    ("1409.0473", "alignment"),
    ("1412.6980", "stochastic"),
    ("1404.5997", "deep convolutional"),
    ("1502.03167", "batch normalization"),
    ("1611.09326", "densely connected"),
    ("2010.11929", "transformer"),
    ("2201.11903", "reasoning"),
    ("2203.02155", "instruction"),
    ("2302.13971", "foundation model"),
]


def build_arxiv_tasks():
    phrasings = [
        "Look up arXiv paper {aid} and return its title and abstract",
        "Get the arXiv paper with ID {aid}",
        "Fetch the abstract for arXiv:{aid}",
        "Download metadata for arXiv {aid}",
    ]
    out = []
    for aid, expected in ARXIV_PAPERS:
        task = phrasings[len(out) % len(phrasings)].format(aid=aid)
        assert_not_leaking(task, expected, aid)
        out.append({
            "task": task,
            "expected_answer": expected,
            "expected_category": "arxiv",
            "id": f"arxiv-{len(out):02d}",
            "artifact": aid,
        })
    return out


# ======================================================================

def main():
    cache = {}

    print("Fetching web content for verification ...", flush=True)
    tasks = []
    tasks += build_excel_tasks()
    print(f"  excel: {len(tasks)}", flush=True)

    prev = len(tasks)
    tasks += build_pdf_tasks()
    print(f"  pdf: {len(tasks) - prev}", flush=True)

    prev = len(tasks)
    tasks += build_filesystem_tasks()
    print(f"  filesystem: {len(tasks) - prev}", flush=True)

    prev = len(tasks)
    tasks += build_arxiv_tasks()
    print(f"  arxiv: {len(tasks) - prev}", flush=True)

    prev = len(tasks)
    tasks += build_wikipedia_tasks(cache)
    print(f"  wikipedia: {len(tasks) - prev}", flush=True)

    prev = len(tasks)
    tasks += build_fetch_tasks(cache)
    print(f"  fetch: {len(tasks) - prev}", flush=True)

    prev = len(tasks)
    tasks += build_search_tasks()
    print(f"  search: {len(tasks) - prev}", flush=True)

    from collections import Counter
    c = Counter(t["expected_category"] for t in tasks)
    print(f"\nTotal: {len(tasks)} tasks")
    for cat, n in sorted(c.items()):
        print(f"  {cat}: {n}")

    # Sanity audit: any tasks with expected leaked in task/artifact?
    leaks = 0
    for t in tasks:
        a = str(t.get("artifact", ""))
        if t["expected_answer"].lower() in t["task"].lower() or \
           t["expected_answer"].lower() in a.lower():
            leaks += 1
            print(f"  LEAK: {t['id']}  expected={t['expected_answer']}")
    print(f"\nLeak audit: {leaks} leaks detected (should be 0)")
    assert leaks == 0, "benchmark has leaks"

    out_path = ROOT / "data" / "gaia_gt.json"
    out_path.write_text(json.dumps(tasks, indent=2))
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
